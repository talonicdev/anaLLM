import numpy as np
import openpyxl
import re
import openpyxl.worksheet
import openpyxl.worksheet.worksheet
import pandas as pd
import json
import csv
import inspect
from pathlib import Path
from dateutil.parser import parse
from sklearn.impute import SimpleImputer
import base64
from io import BytesIO
from enum import Enum
from typing import Optional, Union, Literal, List
import scipy.stats as stats

class LOGEVENT(Enum):
    FILL_EMPTY_CELL = "fill_empty_cells"
    STRIP_TEXT = "strip_text"
    STANDARDIZE_CASE = "standardize_case"
    CONVERT_TO_NUMERIC = "convert_to_numeric"
    REMOVE_SPECIAL_CHARACTERS = "remove_special_characters"
    PARSE_DATES = "parse_dates"
    # ...
    
class DATA_TYPE(Enum):
    INT = "int"                     # Pure integer
    FLOAT = "float"                 # Floating point number
    FORMATTED_NUM = "formatted_num" # Numeric value formatted US/UK style (e.g. "1,000,000")
    DATE = "date"                   # A parseable date
    UNIT = "unit"                   # Unit of measurement
    VALUE_UNIT = "value_unit"       # A numeric value followed by a unit (e.g. "10 km")
    UNIT_VALUE = "unit_value"       # A unit followed by a numeric value (e.g. "$10")
    BOOL = "bool"                   # Boolean value
    WORD = "word"                   # A short string without whitespaces and linebreaks
    STRING = "string"               # Any text that can't be safely interpreted as another data type
    MIXED = "mixed"                 # Seemingly multiple special types; treated as "string", but indicates a need for further user input
    COMPLEX = "complex"             # Numpy type that can't be expressed as a simple type (e.g. ndarray)
    UNKNOWN = "unknown"             # Too few values for a reliable analysis
    
UNITS = [
    # Distance
    "mm","cm","m","km","in","ft","mi","yd","'",'"',
    # Liquids
    "tbsp","cups","ml","l","fl","oz",
    # Currency (ISO)
    "USD","AUD","GBP","EUR","INR","CNY",
    # Currency (Symbol)
    "AU$","$","£","€","₹","¥","ct",
    # Currency (word)
    "Dollars","Euros","Cents","Pounds","Rupees","Yuan","Renminbi",
    # Temperature
    "°F","°C","K","Fahrenheit","Celsius"
    # Time
    "ms","s","min","hr","h","d","seconds","minutes","hours","days","weeks","years",
    # Mass
    "mg","g","kg","t","lb","oz","ton",
    # Movement & Velocity
    "kph","km/h","kn","mph","m/s","mi/s",
    # Data Volumes
    "tb","gb","mb","kb","b","tbit","gbit","mbit","kbit","byte","kilobyte","megabyte",
    # Area
    "mm2","cm2","m2","km2","mi2","ft2","yd2","mm²","cm²","m²","km²","mi²","ft²","yd²",
    # Cubic Area
    "mm3","cm3","m3","km3","mi3","ft3","yd3","mm³","cm³","m³","km³","mi³","ft³","yd³"
]

class DataCleaner:
    def __init__(self, 
                 source: Union[str,pd.DataFrame], 
                 format: Literal['csv','excel','json','base64','dataframe'], 
                 is_structured: Optional[bool] = False, 
                 graceful: Optional[bool] = False,
                 config: Optional[dict] = None
            ):
        """
        Initializes the DataCleaner with the specified data, format, and optional cleaning configurations.

        Parameters:
        :param source (varied): The raw data input specifying the data to work with
            Options:
            - file path to a csv file
            - file path to an excel .xlsx file
            - a JSON string formatted in any of the supported ways (see DataLoaders.from_json)
            - a base64-encoded buffer of an excel .xlsx file
            - a pandas dataframe
        :param format (str): The format of the input data ('csv', 'excel', 'json', 'base64', 'dataframe') which determines how the data is loaded.
        :param is_structured (bool, optional): Indicate that the data to be loaded is guaranteed to be structured, thus skipping checks and transformations.
        :param graceful (bool, optional): If True, illegal or invalid cleaning operations do not raise any exceptions.
        :param config (dict, optional): A dictionary containing configuration options for the cleaning process that specify which operations to perform, customized rules, and other preferences.

        Attributes:
        - source (str, DataFrame): The input data passed to the constructor
        - df (Pandas DataFrame): Loaded data after structure is ensured, working dataframe for all further cleaning methods
        - format (str): A string indicating the format of the incoming data.
        
        - _source_data (Pandas DataFrame or OpenPyXl Worksheet): Loaded data before structure is ensured.
        - _config (dict): Configuration options for customizing the cleaning process.
        - _is_structured (bool): Flag indicating whether or not the data has a recognized structure with headers and consistent columns.
        - _graceful (bool): Flag indicating whether or not failed cleaning operations raise exceptions
        - _cleaned_steps (set): A set of cleaning steps that have already been applied to avoid redundancy.
        - _insignificant_rows (list): A list of strings containing content extracted from removed insignificant rows
        - _changes (list): A list of dicts keeping track of changes affecting the data
        
        - loader (DataLoaders): Responsible for loading data based on the specified format.
        - cleaner (DataCleaners): Contains methods for applying specific cleaning tasks to the data.
        - getter (DataGetters): Provides methods for exporting the cleaned data in various formats.
        - helpers (DataHelpers): Utility functions to support cleaning operations without modifying the data directly.
        """
        
        self.source = source
        self.df: Union[pd.DataFrame,None] = None
        self.format = format
        
        self._source_data: Union[pd.DataFrame,openpyxl.worksheet.worksheet.Worksheet,None] = None
        self._config = config or {}
        self._is_structured = is_structured or False
        self._graceful = graceful or False
        self._cleaned_steps = set()
        self._insignificant_rows: list = []
        self._changes: list = []
        
        self.loader = DataLoaders(self)
        self.cleaner = DataCleaners(self)
        self.getter = DataGetters(self)
        self.detector = DataDetectors(self)
        self.helpers = DataHelpers()
        self.load_data()
        

    def load_data(self):
        load_method = getattr(self.loader, f"from_{self.format}", None)
        if callable(load_method):
            load_method(self.source)
        else:
            raise ValueError(f"Unsupported format: {self.format}")
        
    def is_source_structured(self) -> bool:
        """
        Determines whether or not a given dataframe is structured despite being loaded from a potentially unstructured source.
        """
        if self._is_structured:
            return True
        # TODO: Check data structure for consistency and integrity such that it can be successfully loaded into a Pandas DataFrame, return True if yes
        return False
    
    def structurize_data(self):
        """
        Ensures that the currently held data is structured and performs necessary cleaning actions.
        Populates `df` property, removes meta information rows, names missing headers, renames duplicate columns
        """
        # ...
        if self._is_structured:
            # source data is already structured
            self.df = self._source_data
        else:
            if self._is_structured:
                self._is_structured = True
                # TODO: Create DataFrame from worksheet, populate self.df
                self.df = pd.DataFrame(self._source_data)
            else:
                df = pd.DataFrame()
                # TODO: structurize data and populate self.df
                self._is_structured = True
                self.df = df

    def clean(self, steps=None):
        """
        Applies a series of cleaning steps to the data, if specified, or all available methods.
        Cleaning steps can be configured through the config parameter.

        :param steps: Optional list of cleaning steps to apply. If None, uses config or applies all steps.
        """
        if not self._is_structured:
            self.structurize_data()
            
        steps = steps or self._config.get('enabled_steps', self.cleaner.available_methods())
        for step in steps:
            if step not in self._cleaned_steps and self._config.get(step, {}).get('enabled', True):
                clean_method = getattr(self.cleaner, step, None)
                if callable(clean_method):
                    clean_method(**self._config.get(step, {}))
                    self._cleaned_steps.add(step)
                    
    def _raise(self,exception:Exception):
        """
        Re-raise an exception from within one of the cleaning methods or log the error in case `graceful` is True.

        :param exception: Exception thrown and caught from within DataCleaners.
        """
        if self.graceful:
            frame = inspect.stack()[1].frame
            if 'self' in frame.f_locals:
                # stack item frame has a 'self', so we can assume it's a (sub)class
                # ctx is now `<ClassName>.<FunctionName>`
                ctx = f"{frame.f_locals['self'].__class__.__name__}.{frame.f_code.co_name}"
            else:
                # No 'self' = no class
                # ctx is `<FunctionName>`
                ctx = frame.f_code.co_name
            print(f"Error while executing {ctx}: {exception}")
        else:
            raise exception

    def __apply(self, df_modified, column=None, event_name=''):
        # Compare original data to the modified
        if column:
            original_data = self.df[column]
            modified_data = df_modified[column]
        else:
            original_data = self.df
            modified_data = df_modified

        # Detect changes
        changes = original_data != modified_data
        cells_affected = changes.sum().sum() if isinstance(changes, pd.DataFrame) else changes.sum()

        # Update the DataFrame
        if column:
            self.df[column] = modified_data
        else:
            self.df = modified_data

        # Log the event
        if cells_affected > 0:
            self.log_event({
                "event": event_name,
                "axis": "column" if column else "dataframe",
                "index": column if column else "all_data",
                "cells_affected": int(cells_affected)
            })

    #def log_event(self, log_entry):
    def log_event(self, event: LOGEVENT, axis:str="column", index:str="all", n_cells_affected:int=0):
        # Add a log entry to the logs
        log = {
            "event": event,
            "message": "",
            "axis": axis,
            "index": index,
            "cells_affected": n_cells_affected
        }
        self._changes.append(log)
        

class DataLoaders:
    def __init__(self, main: DataCleaner):
        self.main = main

    def from_csv(self, file_path: str, delimiter=',', encoding='utf-8', quotechar='"', **kwargs):
        """
        Loads data from a CSV file, with parameters for customization that are compatible
        with both pandas.read_csv and csv.reader.

        :param file_path: path to the CSV file.
        :param delimiter: field delimiter (default ',').
        :param encoding: encoding/charset of the file (default 'utf-8').
        :param quotechar: character used to quote fields (default '"').
        :param kwargs: additional keyword arguments specifically for pd.read_csv when structured.
        """
        try:
            if self.main._is_structured:
                # Use pandas to read CSV with all provided kwargs plus defaults
                self.main._source_data = pd.read_csv(
                    file_path, 
                    delimiter=delimiter,
                    encoding=encoding,
                    quotechar=quotechar,
                    **kwargs
                )
            else:
                # Use Python's CSV module to load data into an openpyxl workbook
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                with open(file_path, mode='r', encoding=encoding) as file:
                    reader = csv.reader(file, delimiter=delimiter, quotechar=quotechar)
                    for row_idx, row in enumerate(reader, start=1):
                        for col_idx, value in enumerate(row, start=1):
                            sheet.cell(row=row_idx, column=col_idx, value=value)
                self.main._source_data = sheet
        except Exception as e:
            raise ValueError(f"Failed to load data from CSV: {e}")

    def from_excel(self, file_path: str):
        """
        Loads data from an Excel file.

        :param file_path: Path to the Excel file.
        """
        try:
            if self.main._is_structured:
                self.main._source_data = pd.read_excel(file_path)
            else:
                workbook = openpyxl.load_workbook(file_path)
                self.main._source_data = workbook.active
        except Exception as e:
            raise ValueError(f"Failed to load data from Excel: {e}")

    def from_json(self, json_str: str, orient='index'):
        """
        Converts JSON data into a structured pandas DataFrame if flagged as structured,
        or loads into an openpyxl workbook based on specified format.

        :param json_str: String containing JSON data.
        :param orient: Format of the JSON string (e.g., 'split', 'records', 'index', 'columns', 'values', 'table').
        """
        try:
            if self.main._is_structured:
                self.main._source_data = pd.read_json(json_str, orient=orient, typ='frame')
            else:
                json_data = json.loads(json_str)
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                if orient == 'split':
                    # Example JSON: {'index': [0, 1], 'columns': ['A', 'B'], 'data': [[1, 2], [3, 4]]}
                    columns = json_data['columns']
                    for row_idx, row_data in enumerate(json_data['data'], start=1):
                        for col_idx, value in enumerate(row_data, start=1):
                            sheet.cell(row=row_idx, column=col_idx, value=value)
                    sheet.append(columns)  # Appending column headers
                elif orient == 'records':
                    # Example JSON: [{'A': 1, 'B': 2}, {'A': 3, 'B': 4}]
                    columns = list(json_data[0].keys()) if json_data else []
                    sheet.append(columns)
                    for row_data in json_data:
                        row = [row_data.get(col, None) for col in columns]
                        sheet.append(row)
                elif orient == 'index':
                    # Example JSON: {'0': {'A': 1, 'B': 2}, '1': {'A': 3, 'B': 4}}
                    columns = list(next(iter(json_data.values())).keys()) if json_data else []
                    sheet.append(columns)
                    for index, row_data in sorted(json_data.items()):
                        row = [row_data.get(col, None) for col in columns]
                        sheet.append(row)
                elif orient == 'columns':
                    # Example JSON: {'A': {'0': 1, '1': 3}, 'B': {'0': 2, '1': 4}}
                    columns = list(json_data.keys())
                    sheet.append(columns)
                    row_count = max(len(col_data) for col_data in json_data.values())
                    for i in range(row_count):
                        row = [json_data[col].get(str(i), None) for col in columns]
                        sheet.append(row)
                elif orient == 'values':
                    # Example JSON: [[1, 2], [3, 4]]
                    for row_data in json_data:
                        sheet.append(row_data)
                elif orient == 'table':
                    # Example JSON: {'schema': {'fields': [{'name': 'A'}, {'name': 'B'}]}, 'data': [{'A': 1, 'B': 2}, {'A': 3, 'B': 4}]}
                    columns = [field['name'] for field in json_data['schema']['fields']]
                    sheet.append(columns)
                    for row_data in json_data['data']:
                        row = [row_data.get(col, None) for col in columns]
                        sheet.append(row)
                else:
                    raise ValueError(f"Unsupported JSON format for unstructured data: {orient}")
                self.main._source_data = sheet
        except Exception as e:
            raise ValueError(f"Failed to load data from JSON: {e}")

    def from_base64(self, base64_str: str):
        """
        Loads data from a base64 encoded string of an Excel file.

        :param base64_str: Base64 encoded string of an Excel file.
        """
        try:
            array_buffer = base64.b64decode(base64_str)
            bytes_object = BytesIO(array_buffer)
            if self.main._is_structured:
                self.main._source_data = pd.read_excel(bytes_object)
            else:
                workbook = openpyxl.load_workbook(bytes_object)
                self.main._source_data = workbook.active
        except Exception as e:
            raise ValueError(f"Failed to load data from base64 encoded Excel: {e}")
        
    def from_dataframe(self, dataframe: pd.DataFrame):
        """
        Loads data from an existing Pandas dataframe.

        :param dataframe: The Pandas dataframe.
        """
        self.main._source_data = dataframe
        self.main._is_structured = True
    
class DataGetters:
    def __init__(self, main: DataCleaner):
        self.main = main

    def to_dataframe(self) -> pd.DataFrame:
        """
        Exports the data to a structured pandas DataFrame. Performs necessary actions to ensure that data is structured before exporting.

        :return: Pandas DataFrame containing the structured data.
        """
        if not self.main._is_structured:
            self.main.structurize_data()
        return self.main.df

    def to_json(self, orient="index") -> str:
        """
        Exports the cleaned data to a JSON string. Performs necessary actions to ensure that data is structured for some orients before exporting.
        
        :param orient: JSON string output orientation and format. 
            Allowed values are:
            'split' : dict like {'index' -> [index], 'columns' -> [columns], 'data' -> [values]}
            'records' : list like [{column -> value}, … , {column -> value}]
            'index' : dict like {index -> {column -> value}}
            'columns' : dict like {column -> {index -> value}}
            'values' : just the values array
            'table' : dict like {'schema': {schema}, 'data': {data}}
            
        :return: String containing the JSON data.
        """
        if not self.main._is_structured:
            self.main.structurize_data()
        return self.main.df.to_json(path_or_buf=None, orient=orient)

    def to_csv(self, file_path: str, delimiter=",", encoding="utf-8", quotechar='"', **kwargs) -> str:
        """
        Exports the cleaned data to a CSV file.

        :param file_path: file name, relative or absolute path of the location to save the CSV to
        :param delimiter: Field delimiter (default ',').
        :param encoding: Encoding of the file (default 'utf-8').
        :param quotechar: Character used to quote fields (default '"').
        
        :return: Absolute file path of the saved CSV file
        """
        if not self.main._is_structured:
            self.main.structurize_data()
        self.main.df.to_csv(file_path, delimiter=delimiter, encoding=encoding, quotechar=quotechar, **kwargs)
        return file_path

class DataCleaners:
    def __init__(self, main: DataCleaner):
        self.main = main

    def fill_empty_cells(self, column: str=None, value: Union[float,str,int]=np.nan):
        """
        Converts all empty cells in the column or dataframe to the given values, facilitating easier manipulation
        and analysis of missing data.
        
        :param column: The name of the column to replace values in (default None = all)
        :param value: Replacement value (default NaN)
        """
        pass
    
    def strip_text(self, column: str=None):
        """
        Removes leading and trailing whitespaces.
        
        :param column: The name of the column to strip string values in (default None = all)
        """
        pass

    def standardize_case(self, column: str=None, target_case: str=None):
        """
        Standardizes the case of all text in a specified column. This method determines
        the most common text case (upper, lower, title) and converts all text in the
        column to that case.

        :param column: The name of the column to standardize (default None = all)
        :param target_case: Target case to transform values to (default None = auto-detect)
            Possible values:
            - "upper": Uppercase
            - "lower": Lowercase
            - "initial": First letter to uppercase, rest untouched
        """
        # Handle the case when no specific column is provided; standardize all textual columns
        if column is None:
            columns = self.main.df.select_dtypes(include=['object', 'string']).columns
        else:
            columns = [column]

        for col in columns:
            if self.main.df[col].dtype == 'object' or self.main.df[col].dtype == 'string':
                # Determine the dominant case if not specified
                if target_case is None:
                    values = self.main.df[col].dropna().sample(min(100, len(self.main.df[col]))).astype(str)
                    case_counts = {'upper': 0, 'lower': 0, 'title': 0}
                    for value in values:
                        if value.isupper():
                            case_counts['upper'] += 1
                        elif value.islower():
                            case_counts['lower'] += 1
                        elif value.istitle():
                            case_counts['title'] += 1
                    target_case = max(case_counts, key=case_counts.get) if case_counts else 'lower'

                # Apply the determined case
                if target_case == 'upper':
                    self.main.df[col] = self.main.df[col].str.upper()
                elif target_case == 'lower':
                    self.main.df[col] = self.main.df[col].str.lower()
                elif target_case == 'title':
                    self.main.df[col] = self.main.df[col].str.title()
        pass

    def convert_to_numeric(self, column: str=None):
        """
        Attempts to convert the values in a specified column to numeric types (int or float).
        Handles US/UK number formats such as commas in numbers and strings that contain
        numeric values.

        :param column: The name of the column to convert (default None = all)
        """
        pass

    def remove_special_characters(self, column: str=None):
        """
        Removes special characters from all entries in a specified column. This can help
        clean up text data for analysis, especially when special characters are not relevant.

        :param column: The name of the column from which to remove special characters (default None = all)
        """
        pass

    def parse_dates(self, column: str=None):
        """
        Checks all entries in a specified column and converts them to date format if possible.
        This method can identify various date formats and standardize them into a single
        date format for consistency.

        :param column: The name of the column to process for date validation and parsing (default None = all)
        """
        pass

    def handle_numeric_ranges(self, column: str=None):
        """
        Identifies and processes numeric ranges, potentially splitting
        these ranges into separate data points or standardizing their format.

        :param column: The name of the column containing numeric ranges (default None = all)
        """
        pass

    def drop_duplicate_rows(self):
        """
        Removes duplicate rows from the data, helping to ensure that the dataset contains
        only unique entries.
        """
        pass
    
    def drop_empty_rows(self):
        """
        Removes rows containing only falsy values from the data.
        """
        try:
            self.main.df.dropna(how='all', inplace=True)
        except Exception as e:
            self.main._raise(e)
    
    def drop_empty_columns(self):
        """
        Removes columns containing only falsy values from the data.
        """
        try:
            self.main.df.dropna(axis='columns', how='all', inplace=True)
        except Exception as e:
            self.main._raise(e)

    def rename_duplicate_columns(self):
        """
        Identifies columns with duplicate names and renames them to ensure that each column
        has a unique identifier. This is crucial for data manipulation and retrieval operations.
        """
        pass

    def extract_insignificant_rows(self):
        """
        Eliminates rows that do not provide significant information, such as headers or footers
        that are included in the data. Identified rows can be retrieved as a list later.
        """
        pass

    def impute_missing_values(self, column: str=None, strategy: Literal['mean','median','min','max','frequent']='mean'):
        """
        Use statistical methods like mean or median to estimate missing values.

        :param column: The name of the column containing numeric ranges (default None = all)
        :param strategy: Statistical value to fill missing values with (default 'mean')
        """
        pass

    def validate_units(self, column: str=None):
        """
        Verifies and standardizes the units of measurement in a specified column, ensuring
        that all data points in the column are consistent and comparable.

        :param column: The name of the column whose units need validation (default None = all)
        """
        pass
    
    def split_value_unit(self, column: str=None):
        """
        Splits entries in a specified column that contain both numeric values and their units
        into two separate columns: one for the numeric value and one for the unit. This is
        particularly useful for columns where entries are formatted like '100 ft' or '$10', 
        ensuring that numeric operations can be performed on the data while retaining unit 
        information in a separate column.

        :param column: The name of the column to process. The method expects this column to
                       contain strings containing both a numeric value and a unit, e.g. '$10', '20ft', '30 kph' (default None = all)
        """
        # We can extend this later, but for our markets, this should suffice
        units = [
            # Distance
            "mm","cm","m","km","in","ft","mi","'",'"',
            # Liquids
            "tbsp","cups","ml","l","fl","oz",
            # Currency (ISO)
            "USD","AUD","GBP","EUR","INR","CNY",
            # Currency (Symbol)
            "AU$","$","£","€","₹","¥",
            # Temperature
            "°F","°C","K",
            # Time
            "ms","s","min","hr","h","d",
            # Mass
            "mg","g","kg","t","lb","oz","ton",
            # Movement & Velocity
            "kph","km/h","kn","mph","m/s","mi/s"
        ]
        pass

class DataHelpers:
    def is_range(self, string: str) -> bool:
        """
        Determines if the given string represents a numeric range, typically formatted
        as 'start-end' (e.g., '100-200'). This can be useful for identifying range data
        that may need special processing.

        :param string: The string to evaluate.
        :return: True if the string is a numeric range, False otherwise.
        """
        pass

    def infer_text_case(self, values: list) -> str:
        """
        Infers the most common text case (uppercase, lowercase, title case) in a list of
        string values. This method can help determine the dominant text case for
        standardizing text data in a column.

        :param values: A list of string values.
        :return: A string indicating the most common case ('upper', 'lower', 'title').
        """
        pass

    def is_date(self, string: str, fuzzy: bool = False) -> bool:
        """
        Checks whether a given string can be interpreted as a date. This function uses
        various date formats to attempt parsing the string and validates it as a date.

        :param string: The string to check.
        :param fuzzy: If True, allows the parsing to ignore unknown tokens in the string.
        :return: True if the string can be parsed as a date, False otherwise.
        """
        pass

    def is_special_character(self, elem: str) -> bool:
        """
        Identifies if the string contains any special characters. This is useful for cleaning
        operations where special characters need to be removed or flagged.

        :param elem: The string to inspect.
        :return: True if any special characters are found, False otherwise.
        """
        pass

    def check_number_unit(self, elem: str, units_enum) -> bool:
        """
        Checks if any part of the string corresponds to known units from a given enumeration
        of units. This helps validate and standardize units in data entries.

        :param elem: The string containing potential unit designations.
        :param units_enum: An enumeration of accepted units.
        :return: True if the units are valid as per the enumeration, False otherwise.
        """
        pass

class DataDetectors:
    def __init__(self, main: DataCleaner):
        self.main = main
        
    @staticmethod
    def _calc_sample_size(population:int,confidence_level:float,margin_of_error:float) -> int:
        Z = stats.norm.ppf(1 - (1 - confidence_level) / 2)
        p = 0.5
        n0 = (Z**2 * p * (1 - p)) / (margin_of_error**2)
        return int(n0 / (1 + ((n0 - 1) / population)))
    
    def determine_column_type(self, column: str, accuracy: float = 0.95) -> DATA_TYPE:
        """
        Inspect a given {column} and try to determine - with reasonable certainty - the data type common to the column, if any.
        
        Allows for missing (Null, NaN, "") values and a small amount of outliers.

        :param column: The name of the column to determine the data type of
        :param accuracy: Desired confidence level, inverse margin of error and maximum proportion of outliers (Default 0.95)
        
        :return: Data type literal.
        """
        
        # Get a list of potential data types from a small set of samples
        peek_sample_size = 10
        peek_samples = self.get_sample_values(column, count=peek_sample_size)
        potential_types = []
        for value in peek_samples:
            detected_types = self.infer_data_types(value)
            potential_types.extend(detected_types)
        unique_types = list(set(potential_types))

        # Get a list of sample values that ensures minimum accuracy
        population = len(self.main.df[column])
        margin_of_error = 1 - accuracy
        sample_size = self._calc_sample_size(
            population=population,
            confidence_level=accuracy,
            margin_of_error=margin_of_error)
        samples = self.get_sample_values(column, count=sample_size)
        
        # If there are less samples than necessary, return UNKNOWN data type
        if len(samples) < sample_size:
            return DATA_TYPE.UNKNOWN
        
        type_counts = {data_type: 0 for data_type in unique_types}

        # Count possible data types for all sample values
        for value in samples:
            detected_types = self.infer_data_types(value)
            for data_type in unique_types:
                if data_type in detected_types:
                    type_counts[data_type] += 1

        # Determine the most frequent data type
        total_samples = len(samples)
        for data_type, count in type_counts.items():
            if count / total_samples >= accuracy:
                return data_type

        return DATA_TYPE.UNKNOWN
        
        
    def get_sample_values(self, column: str, count: int = 3, variance_ratio: float = 0.2) -> list[str]:
        """
        Retrieve a list of {count} existing values from a given {column},
        selected from across the entire column and with some randomized variance.
        
        Sampled value are returned stripped and stringified.

        :param column: The name of the column to select values from
        :param count: The amount of values to retrieve
        :param variance_ratio: maximum distance from the non-randomized sample
        
        :return: List of strings.
        """
        column_values = self.main.df[column].dropna().values
        num_values = len(column_values)

        if count >= num_values:
            unique_values = np.unique(column_values)
            safe_values = []
            for value in unique_values:
                try:
                    value_str = str(value).strip()
                    if value_str:
                        safe_values.append(value_str)
                except:
                    pass
            return safe_values

        results = []
        sampled_indices = set()
        attempts = 0
        max_attempts = count * 10  # Limiting attempts to avoid potential infinite loops

        while len(results) < count and attempts < max_attempts:
            base_index = np.random.randint(num_values)
            variance = int(variance_ratio * num_values)
            start_index = max(0, base_index - variance)
            end_index = min(num_values - 1, base_index + variance)

            sample_index = np.random.randint(start_index, end_index + 1)

            if sample_index not in sampled_indices:
                value = column_values[sample_index]
                try:
                    value_str = str(value).strip()
                    if value_str:
                        results.append(value_str)
                        sampled_indices.add(sample_index)
                except:
                    pass
            attempts += 1

        return results
    
    @staticmethod
    def infer_data_types(value: str) -> List[DATA_TYPE]:
        """
        Infer possible data types for a given value.

        :param value: The value to classify.
        :return: A list of DATA_TYPE enums that could potentially describe the value.
        """
        potential_types = []
        
        if DataDetectors.is_complex_type(value):
            return [DATA_TYPE.COMPLEX]

        if DataDetectors.is_numeric(value):
            if DataDetectors.is_int(value):
                potential_types.append(DATA_TYPE.INT)
            if DataDetectors.is_float(value):
                potential_types.append(DATA_TYPE.FLOAT)
            if DataDetectors.is_formatted_number(value):
                potential_types.append(DATA_TYPE.FORMATTED_NUM)
        if DataDetectors.is_date(value):
            potential_types.append(DATA_TYPE.DATE)
        if DataDetectors.is_unit(value):
            potential_types.append(DATA_TYPE.UNIT)
        if DataDetectors.is_value_unit(value):
            potential_types.append(DATA_TYPE.VALUE_UNIT)
        if DataDetectors.is_unit_value(value):
            potential_types.append(DATA_TYPE.UNIT_VALUE)
        if DataDetectors.is_bool(value):
            potential_types.append(DATA_TYPE.BOOL)
        if DataDetectors.is_single_word(value):
            potential_types.append(DATA_TYPE.WORD)
        if not potential_types:  # If no specific type matched
            potential_types.append(DATA_TYPE.STRING)  # Consider it a STRING by default

        return potential_types
    
    @staticmethod
    def is_simple_type(value):
        """
        Checks if the value is a simple type, including native Python types and their numpy equivalents that are typically used in 2D tables.
        """
        simple_types = (int, float, str, bool, np.integer, np.float_, np.bool_, np.int_)
        return isinstance(value, simple_types)
    
    @staticmethod
    def is_complex_type(value):
        """
        Determines if the value is considered complex, meaning that it is not a value easily expressed as a string.
        """
        return not DataDetectors.is_simple_type(value)
    
        
    @staticmethod
    def is_numeric(value:str) -> bool:
        """
        Determines whether or not a given value appears to be a numeric value of any kind (int, float, any numpy numeric type, scientific or casual notation with thousands separators).
        """
        try:
            return bool(re.search(f"^-?[0-9e,. ]+$",value))
        except:
            return False
    
    @staticmethod
    def is_int(value:str) -> bool:
        """
        Determines whether or not a given value appears to be a positive or negative integer.
        """
        try:
            int(value)
            return True
        except ValueError:
            return False
        
    @staticmethod
    def is_float(value:str) -> bool:
        """
        Determines whether or not a given value appears to be a floating point variable (incl np floats).
        """
        try:
            float(value)
            return '.' in value or 'e' in value.lower()  # Ensure it's a floating-point representation
        except ValueError:
            return False
                
    @staticmethod
    def is_formatted_number(value:str) -> bool:
        """
        Determines whether or not a given value appears to be an integer with US/UK style thousands separators.
        """
        try:
            if re.match(r'^-?[0-9]{1,3}(,[0-9]{3})*(\.[0-9]+)?$', value):
                return True
            return False
        except:
            return False
    
    @staticmethod
    def is_unit(value:str) -> bool:
        """
        Determines whether or not a given value appears to be a unit of measurement.
        """
        return value in UNITS
        
    @staticmethod
    def is_bool(value:str) -> bool:
        """
        Determines whether or not a given value appears to indicate a boolean value (True, False).
        """
        bool_patterns = [
            r'^true$',
            r'^false$',
            r'^yes$',
            r'^no$',
            r'^1$',
            r'^0$',
            r'^y$',
            r'^n$'
        ]
        try:
            val_lower = str(value).lower()
            return bool(any(re.match(pattern, val_lower) for pattern in bool_patterns))
        except:
            return False
    
    @staticmethod
    def is_single_word(value:str) -> bool:
        """
        Determines whether or not a given value appears to be a short string (<= 100 chars) without whitespace or line breaks.
        """
        try:
            return bool(re.search(r"^[^\s]{1,100}$",str(value)))
        except:
            return False
        
    @staticmethod
    def is_date(value:str) -> bool:
        """
        Determines whether or not a given value appears to be a date.
        """
        date_patterns = [
            r'\d{4}-\d{1,2}-\d{1,2}',  # Matches YYYY-MM-DD
            r'\d{1,2}/\d{1,2}/\d{4}',  # Matches MM/DD/YYYY
            r'\d{1,2} \w+ \d{4}',      # Matches DD Month YYYY
            r'\w+ \d{1,2}, \d{4}',      # Matches Month DD, YYYY
            r'\d{1,2}-\d{1,2}-\d{4}',  # Matches DD-MM-YYYY
            r'\d{1,2}\.\d{1,2}\.\d{4}', # Matches DD.MM.YYYY
            r'\d{4}/\d{1,2}/\d{1,2}',  # Matches YYYY/DD/MM
            r'\d{4}-\d{1,2}-\d{1,2} \d{2}:\d{2}(:\d{2})?', # Matches YYYY-MM-DD HH:MM(:SS)
            r'\w{3}, \d{2} \w{3} \d{4} \d{2}:\d{2}:\d{2} GMT', # Matches Day, DD Mon YYYY HH:MM:SS GMT
            r'\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}\.\d{3}Z', # Matches YYYY-MM-DDTHH:MM:SS.MSZ
            r'\d{1,2}:\d{2}:\d{2} (AM|PM)' # Matches HH:MM:SS AM/PM
        ]
        # Check if the string matches any of the date patterns
        try:
            if any(re.match(pattern, value) for pattern in date_patterns):
                try:
                    parse(value, fuzzy=True)
                    return True
                except ValueError:
                    return False
            else:
                return False
        except:
            return False
        
    @staticmethod
    def is_value_unit(value:str) -> bool:
        """
        Determines whether or not a given value appears to be a value followed by a unit of measurement (e.g. "10 km").
        """
        try:
            pattern = r'^-?\d+(\.\d+)?\s*(' + '|'.join(re.escape(unit) for unit in UNITS) + ')$'
            return bool(re.match(pattern, value))
        except:
            return False
    
    @staticmethod
    def is_unit_value(value:str) -> bool:
        """
        Determines whether or not a given value appears to be a unit of measurement followed by a value (e.g. "$10").
        """
        try:
            pattern = r'^(' + '|'.join(re.escape(unit) for unit in UNITS) + ')\s*-?\d+(\.\d+)?$'
            return bool(re.match(pattern, value))
        except:
            return False