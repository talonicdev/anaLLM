from enum import Enum

import numpy as np
from dateutil.parser import parse

import yaml
import openpyxl
import re
import os
import openai
import pandas as pd
import base64
from io import BytesIO
from typing import Optional

from pathlib import Path
from decouple import config

from langchain_core.prompts import FewShotChatMessagePromptTemplate, ChatPromptTemplate


class ExtendedEnum(Enum):
    @classmethod
    def abbreviations(cls):
        return list(map(lambda c: re.findall(r'\(.*?\)', c.value)[0][1:-1].lower(), cls))

    @classmethod
    def names(cls):
        return list(map(lambda c: c.value.split(' (')[0].lower(), cls))


class Units(ExtendedEnum):
    pass


class WordContext(Enum):
    pass


class WordIndustries(Enum):
    pass


class WordException(Enum):
    MANAGEMENT = "management"
    ROW = "row"
    COLUMN = "column"
    DATA = "data"
    ANALYSIS = "analysis"
    ANALYTICS = "analytics"

def init_workbook_from_json(json):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    # Get max indices so we don't accidentally skip any content
    max_row_index = max(int(key) for key in json.keys())
    max_col_index = max(max(int(key) for key in row.keys()) for row in json.values() if row)

    for row_index in range(max_row_index + 1):
        row_key = str(row_index)
        for col_index in range(max_col_index + 1):
            col_key = str(col_index)
            value = json.get(row_key, {}).get(col_key, '')
            cell = sheet.cell(row=row_index + 1, column=col_index + 1)
            cell.value = value

    return workbook

def init_workbook_from_base64XLSXbuffer(base64string: str):
    array_buffer = base64.b64decode(base64string)
    bytes_object = BytesIO(array_buffer)
    return openpyxl.load_workbook(bytes_object)
    
def init_workbook_from_file(file_path: str):
    return openpyxl.load_workbook(file_path)

def convert_to_numeric(value) -> int|float:
    if isinstance(value, str):
        # Check for any characters that are not digits, commas, periods, or space
        if re.search(r'[^\d,.\s]', value):
            raise ValueError()

        # Remove all spaces, leaving only digits, commas and periods
        cleaned_value = re.sub(r'\s+', '', value)

        # Split into parts before and after the decimal point
        if '.' in cleaned_value:
            integer_part, decimal_part = cleaned_value.split('.', 1)
        else:
            integer_part, decimal_part = cleaned_value, None

        # Validate comma placement for thousand separators in the integer part
        if ',' in integer_part:
            parts = integer_part.split(',')
            if len(parts[0]) > 3 or any(len(part) != 3 for part in parts[1:]):
                # Incorrect comma placement
                raise ValueError()  
            # Remove commas now that they're verified
            integer_part = integer_part.replace(',', '')
        
        # Reassemble the number if there's a decimal part
        if decimal_part is not None:
            cleaned_value = f"{integer_part}.{decimal_part}"
        else:
            cleaned_value = integer_part

        # Convert to number, ensuring it's a valid numeric format
        num = float(cleaned_value)
        return int(num) if num.is_integer() else num

    return value

def infer_text_case(values):
    case_counts = {'upper': 0, 'lower': 0, 'title': 0}
    for value in values:
        if isinstance(value, str):
            if value.isupper():
                case_counts['upper'] += 1
            elif value.islower():
                case_counts['lower'] += 1
            elif value.istitle():
                case_counts['title'] += 1

    return max(case_counts, key=case_counts.get) if case_counts else 'lower'

def standardize_text_columns(df, min_uniformity_threshold=0.8):
    for column in df.columns:
        # Continue if the column is purely numeric
        if df[column].dtype in ['int64', 'float64']:
            continue

        # Check if the column contains text
        if df[column].apply(lambda x: isinstance(x, str)).all():
            # Convert to string and strip whitespace
            df[column] = df[column].astype(str).str.strip()

            # Sample for case uniformity analysis
            sample_values = df[column].dropna().head(20)
            if not sample_values.empty:
                dominant_case = infer_text_case(sample_values)
                uniformity = sample_values.apply(lambda x: x.isupper() if dominant_case == 'upper' else 
                                                             x.islower() if dominant_case == 'lower' else 
                                                             x.istitle()).mean()

                # Apply standardization if uniformity is high
                if uniformity >= min_uniformity_threshold:
                    if dominant_case == 'upper':
                        df[column] = df[column].str.upper()
                    elif dominant_case == 'lower':
                        df[column] = df[column].str.lower()
                    elif dominant_case == 'title':
                        df[column] = df[column].str.title()

    return df

def rename_duplicate_columns(columns):
    counts = {}
    new_columns = []
    for col in columns:
        if col in counts:
            counts[col] += 1
            new_columns.append(f"{col} ({counts[col] + 1})")
        else:
            counts[col] = 1
            new_columns.append(col)
    return new_columns

def format_dataframe(data, format: Optional[str]='file', table_name: Optional[str]='Untitled Table - Untitled'):
    
    # Load the workbook from the specified format
    if format == 'file':
        # Load from file -> `data` is a file path
        workbook = init_workbook_from_file(data)
        table_name = Path(data).stem
    elif format == 'json':
        # Load from JSON -> `data` is { [row index: number]: {col index: number}: string|number }
        workbook = init_workbook_from_json(data)
    elif format == 'excel':
        # Load from Excel -> `data` is a base64 encoded .xlsx file buffer
        workbook = init_workbook_from_base64XLSXbuffer(data)
    else:
        # Try loading from some other format, presumably dict, skip formatting/cleaning
        return pd.DataFrame(data), [table_name]
    
    worksheet = workbook.active
    row_vectors = []

    # Process each row in the worksheet to determine significant rows
    for row in worksheet.iter_rows(1, worksheet.max_row):
        # Append non-empty cells as tuples (index, cell_value)
        row_vectors.append([i for i, cell in enumerate(row) if cell.value])

    # Determine the length of each row vector
    row_lens = [len(row) for row in row_vectors]
    
    # Find the maximum row length
    max_row_len = np.max(row_lens)
    
    # Keep rows that have significant content (more than half of the max length)
    keep_idx = [i for i in range(len(row_lens)) if row_lens[i] > max_row_len//2]

    data = []
    col_names = []
    side_data = [table_name]

    if keep_idx:
        # Extract column names and row data
        for idx, row in enumerate(worksheet.iter_rows(1, worksheet.max_row)):
            if idx in keep_idx:
                if idx == keep_idx[0]:
                    # First significant row is treated as the header
                    col_names = [cell.value for cell in row]
                    # Rename duplicate columns Windows style ("Name", "Name (2)", "Name (3)", ..)
                    col_names = rename_duplicate_columns(col_names)
                else:
                    # Subsequent rows are treated as data
                    vals = {col_names[i]: cell.value for i, cell in enumerate(row)}
                    data.append(vals)
            else:
                # Rows before the first significant row are treated as side data
                if idx < keep_idx[0]:
                    vals = [cell.value for cell in row if cell.value]
                    if len(vals) > 0:
                        side_data.extend([cell.value for cell in row if cell.value])
                        
        if len(keep_idx) == 1:  # This means only header was found, no data rows
            df = pd.DataFrame(columns=col_names)
        else:
            df = pd.DataFrame(data, columns=col_names)
                        
    else:
        df = pd.DataFrame()

    # Create DataFrame from data
    #df = pd.DataFrame.from_records(data)
    print(df.columns)
    # Remove empty rows
    df = df.dropna(how='all')
    # Remove empty columns
    #df = df.dropna(axis='columns', how='all')
    # Fill missing cell values with NaN
    df = df.fillna(value=np.nan)
    # Drop duplicate rows (across all columns, keeping the respective first instance)
    df = df.drop_duplicates()
    # Standardize string cases (upper, lower, initials)
    df = standardize_text_columns(df)
    # Try to convert columns to int/float (US/UK notation only) for all columns with entirely numeric (or NaN) cell values
    for column in df.columns:
        try:
            converted_column = df[column].apply(convert_to_numeric)
            df[column] = converted_column
        except:
            pass

    # If no side data was extracted, use the table name as side data
    #if len(side_data) == 0:
    #    side_data.append(table_name)
        
    #print(df)

    return df, side_data


def load_templates(template_name: str) -> tuple:
    script_path = Path(__file__).parents[1].resolve()

    with open(script_path / f'{template_name}/example_template.txt', 'r') as file:
        template = file.read().replace('\n', ' \n ')

    with open(script_path / f'{template_name}/prefix.txt', 'r') as file:
        prefix = file.read().replace('\n', ' \n ')

    with open(script_path / f'{template_name}/suffix.txt', 'r') as file:
        suffix = file.read()

    with open(script_path / f'{template_name}/examples.yaml', 'r') as file:
        examples = yaml.safe_load(file)

    examples = [examples[k] for k in examples.keys()]
    return template, prefix, suffix, examples


def get_template(examples, prefix, suffix) -> ChatPromptTemplate:
    example_prompt = ChatPromptTemplate.from_messages(
        [('human', suffix), ('ai', '{answer}')]
    )
    few_shot_prompt = FewShotChatMessagePromptTemplate(
        examples=examples,
        example_prompt=example_prompt,
    )
    return ChatPromptTemplate.from_messages([('system', prefix), few_shot_prompt, ('human', suffix)])


def load_datasets(subset: int = -1):
    script_path = Path(__file__).parent.resolve()
    datasets_path = script_path / 'datasets/original'

    file_list = [f for f in datasets_path.glob('**/*') if f.is_file()][:subset]

    table_list = []

    for file in file_list:
        file_extension = file.suffix
        table = pd.read_csv(file) if file_extension == '.csv' else pd.read_excel(file)
        table_list.append(table)

    return table_list


def parse_string(elem: str) -> dict:
    numbers = []
    others = []
    idx = 0

    p = re.compile("([a-zA-Z]+)")
    matches = p.finditer(elem)
    matches = tuple(matches)
    for i, m in enumerate(p.finditer(elem)):
        others.append(m.group())
        if elem[idx: m.start()]:
            numbers.append(elem[idx: m.start()])
        if i == len(matches) - 1:
            if elem[m.end():]:
                numbers.append(elem[m.end():])
        idx = (m.start() + len(m.group()))

    return {'numbers_list': numbers, 'others_list': others}


def parse_number(numbers_list: list, others_list: list):
    separators = ['.', ',']
    seps = []
    unit = None
    for e in others_list:
        if e in separators:
            seps.append(e)
        else:
            unit = e

    if len(seps) > 1:
        float_digit = float(''.join(numbers_list[: -1])) + float('0.' + numbers_list[-1])
    elif len(seps) == 1:
        if seps[0] == ',':
            float_digit = float('.'.join(numbers_list))
        else:
            float_digit = float('.'.join(numbers_list))
    else:
        float_digit = int(numbers_list[0])

    return float_digit, unit


def is_date(string, fuzzy=False):
    """
    Return whether the string can be interpreted as a date.

    :param string: str, string to check for date
    :param fuzzy: bool, ignore unknown tokens in string if True
    """
    date_patterns = [
        r'\d{4}-\d{1,2}-\d{1,2}',  # Matches YYYY-MM-DD
        r'\d{1,2}/\d{1,2}/\d{4}',  # Matches MM/DD/YYYY
        r'\d{1,2} \w+ \d{4}',      # Matches DD Month YYYY
        r'\w+ \d{1,2}, \d{4}',      # Matches Month DD, YYYY
        r'\d{1,2}-\d{1,2}-\d{4}',  # Matches DD-MM-YYYY
        r'\d{1,2}\.\d{1,2}\.\d{4}', # Matches DD.MM.YYYY
        r'\d{4}/\d{1,2}/\d{1,2}',  # Matches YYYY/DD/MM
        r'\d{4}-\d{1,2}-\d{1,2} \d{2}:\d{2}(:\d{2})?', # Matches YYYY-MM-DD HH:MM(:SS)
        r'\w{3}, \d{2} \w{3} \d{4} \d{2}:\d{2}:\d{2} GMT', # Matches Day, DD Mon YYYY HH:MM:SS GMT
        r'\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}\.\d{3}Z', # Matches YYYY-MM-DDTHH:MM:SS.MSZ
        r'\d{1,2}:\d{2}:\d{2} (AM|PM)' # Matches HH:MM:SS AM/PM
    ]
    # Check if the string matches any of the date patterns
    if any(re.match(pattern, string) for pattern in date_patterns):
        try:
            parse(string, fuzzy=fuzzy)
            return True
        except ValueError:
            return False
    else:
        return False


def is_range(string):
    if '-' in string:
        if len(string.split('-')) == 2:
            start, end = string.split('-')
            if len(start) > 0:
                if start.isnumeric() and end.isnumeric():
                    return True
    else:
        return False


def special_char(elem: str):
    special = re.compile("[!@#&*()_+=|<>?{}\\[\\]~]")
    return special.search(elem)


def check_number_unit(elem: str, units):
    res = parse_string(elem)
    if len(res['others_list']) > 0:
        if np.all([x.lower() in units.names() + units.abbreviations() for x in res['others_list']]):
            return True
    else:
        return False


def prep_table(table: pd.DataFrame):
    table = table.dropna()
    table = table.replace([np.inf, -np.inf], 0)
    x = table.infer_objects().dtypes
    #return table
    for i, col in enumerate(table.columns):
        if x[col] == 'object':
            if np.all([isinstance(x, str) for x in table[col].tolist()]):
                if np.all([bool(re.search(r'\d', x)) for x in table[col].tolist()]):
                    if not np.any([is_date(x) for x in table[col].tolist()]):
                        if not np.any([is_range(x) for x in table[col].tolist()]):
                            if not np.any([special_char(x) for x in table[col].tolist()]):
                                if len(parse_string(table[col].iloc[0])['others_list']) < 3:
                                    if check_number_unit(table[col].iloc[0], Units):
                                        val_units_df = table[col].apply(lambda x: parse_number(**parse_string(x)))
                                        df_new = pd.DataFrame.from_records(val_units_df.to_list(),
                                                                           columns=[f'{col}', 'unit'])
                                        table.drop(columns=col, inplace=True)
                                        table = pd.concat([table, df_new], axis=1)

    return table
